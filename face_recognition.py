import cv2
import numpy as np 
import os 
import openpyxl
import datetime
import pandas as pd
import tkinter as tk
from tensorflow import expand_dims
from keras.models import load_model
import matplotlib.pyplot as plt
from mtcnn.mtcnn import MTCNN
from matplotlib.patches import Rectangle
import pickle
import numpy as np

with open('./saved/id_to_roll_f.pkl','rb') as f:
	try:
		id_to_roll = pickle.load(f)
	except EOFError:
		id_to_roll = {}
with open('./saved/roll_to_id_f.pkl','rb') as f:
	try:
		roll_to_id = pickle.load(f)
	except EOFError:
		roll_to_id = {}

model =load_model(
    './saved/model/Facial_recogNet.h5')

def face_recognizer(photo):
	photo = np.expand_dims(photo,axis=0)
	ypreds = model.predict(photo)
	ypred = np.argmax(ypreds) 
	return ypred

detector = MTCNN()
def mark_attendance():
	var = tk.messagebox.showinfo("Notification", "If Your Name Appears on your Face Press esc !")

	#Init Camera
	cap = cv2.VideoCapture(0)
	# Face Detection
	# Data Preparation

	student_table=pd.read_csv('student_details.csv')

	# Testing 
	roll=""
	while True:
		ret,frame = cap.read()
		if ret == False:
			continue

		faces = detector.detect_faces(frame)
		if len(faces)==0:
			print('Your face is not visible , please get into the frame\n')
			continue

		offset = 10
		x, y, w, h  = faces[0]['box']
		cv2.rectangle(frame,(x-offset,y-offset),(x+w+offset,y+h+offset),(0,255,255),2)
		face_section = frame[y-offset:y+h+offset,x-offset:x+w+offset]
		face_section = cv2.resize(face_section,(224,224))
		face_section = cv2.cvtColor(face_section,cv2.COLOR_BGR2RGB)
		#Predicted Label (out)
		out = face_recognizer(face_section)
		#Display on the screen the name and rectangle around it
		roll = id_to_roll.get(out)
		cv2.putText(frame,roll,(x,y-10),cv2.FONT_HERSHEY_SIMPLEX,1,(255,0,0),2,cv2.LINE_AA)
		cv2.imshow("FRAME",frame)
		key = cv2.waitKey(1) & 0xFF
		if key == 27:
			break

	# Marking attendance of student in attendance sheet
	cap.release()
	cv2.destroyAllWindows()

	def load_workbook(wb_path):
		if os.path.exists(wb_path):
			return openpyxl.load_workbook(wb_path)
		return openpyxl.Workbook()
		
	name=student_table[student_table['roll']==roll].iloc[0]['name']
	dt=datetime.datetime.now()
	date = dt.strftime("%m-%d-%Y")
	time = dt.strftime("%H:%M:%S")
	filename=date
	wb_path='Attendance/'+filename+'.xlsx'
	print(wb_path)
	wb=load_workbook(wb_path)
	sheet=wb['Sheet']
	sheet.append((name,roll,time))
	wb.save(wb_path)

	

	tk.messagebox.showinfo("Notification", roll + "! Your attendance is marked") 
